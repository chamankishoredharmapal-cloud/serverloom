# accounts/views.py
from django.contrib.auth.models import User
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, HttpResponseBadRequest
from django.utils import timezone
from django.db.models import Sum, Q
from django.db import transaction

from core.models import (
    Employee, SareeCount, PagdiHistory, SalaryHistory,
    WarpHistory, AdvanceHistory, PagdiChangeHistory
)
from core import services

# openpyxl used for XLSX export
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


# ---------------------------------------------------------
# Helpers & decorators
# ---------------------------------------------------------
def _is_staff(user):
    return bool(user and (user.is_staff or user.is_superuser))


staff_required = user_passes_test(_is_staff, login_url="login")


# =========================================================
# AUTH
# =========================================================
def signup_view(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        phone = request.POST.get("phone", "").strip()
        password = request.POST.get("password", "")

        if not name or not phone or not password:
            return render(request, "accounts/signup.html", {"error": "All fields are required."})

        if User.objects.filter(username=phone).exists():
            return render(request, "accounts/signup.html", {"error": "Phone already registered"})

        user = User.objects.create_user(username=phone, password=password)
        Employee.objects.create(
            user=user,
            name=name,
            phone=phone,
            salary_per_saree=0,
            advance_salary=0,
            is_approved=False
        )

        messages.success(request, "Account created — waiting for admin approval.")
        return redirect("login")

    return render(request, "accounts/signup.html")


def login_view(request):
    if request.method == "POST":
        phone = request.POST.get("phone")
        password = request.POST.get("password")

        user = authenticate(request, username=phone, password=password)

        if not user:
            return render(request, "accounts/login.html", {"error": "Invalid phone or password."})

        emp = Employee.objects.filter(user=user).first()

        if emp:
            if not emp.is_approved:
                return render(request, "accounts/login.html", {"error": "Account not approved yet."})

            login(request, user)
            request.session["employee_id"] = emp.id
            return redirect("employee_dashboard")

        if user.is_staff or user.is_superuser:
            login(request, user)
            request.session["employee_id"] = None
            return redirect("admin_home")

        return render(request, "accounts/login.html", {"error": "Unauthorized account."})

    return render(request, "accounts/login.html")


def logout_view(request):
    logout(request)
    return redirect("login")


# =========================================================
# EMPLOYEE VIEWS
# =========================================================
@login_required
def employee_dashboard(request):
    """
    Employee dashboard. No ability to add saree counts here (read-only).
    """
    emp_id = request.session.get("employee_id")
    if not emp_id:
        messages.error(request, "Session expired or no employee session. Please login again.")
        return redirect("login")

    employee = get_object_or_404(Employee, id=emp_id)

    today = timezone.localdate()
    monday, sunday = services.get_week_bounds(today)

    weekly_sarees = SareeCount.objects.filter(
        employee=employee, date__gte=monday, date__lte=sunday
    ).aggregate(Sum("count"))["count__sum"] or 0

    weekly_salary_before = weekly_sarees * (employee.salary_per_saree or 0)
    final_salary = weekly_salary_before - (employee.advance_salary or 0)

    return render(request, "accounts/dashboard.html", {
        "employee": employee,
        "weekly_sarees": weekly_sarees,
        "sarees": weekly_sarees,
        "salary_per_saree": employee.salary_per_saree,
        "advance": employee.advance_salary,
        "final_salary": final_salary,
        "week_start": monday,
        "week_end": sunday,
    })


@login_required
def saree_count_view(request):
    """
    Employee saree history page. Employees CANNOT add saree counts here.
    Compute salary per row in view to avoid template arithmetic.
    """
    employee = get_object_or_404(Employee, user=request.user)

    rows = SareeCount.objects.filter(employee=employee).order_by("-date").select_related("employee")

    history = []
    rate = employee.salary_per_saree or 0
    for r in rows:
        history.append({
            "id": r.id,
            "date": r.date,
            "count": r.count,
            "notes": r.notes,
            "salary": r.count * rate
        })

    return render(request, "accounts/saree_count.html", {
        "employee": employee,
        "history": history
    })


@login_required
def pagdi_view(request):
    employee = get_object_or_404(Employee, user=request.user)

    active = PagdiHistory.objects.filter(employee=employee, end_date__isnull=True).first()

    if request.method == "POST" and active:
        services.finish_pagdi(active.id, request.user, "Finished by employee")
        messages.success(request, "Pagdi finished.")
        return redirect("pagdi")

    sarees_made = 0
    remaining = 0

    if active:
        sarees_made = SareeCount.objects.filter(employee=employee, date__gte=active.start_date).aggregate(Sum("count"))["count__sum"] or 0
        remaining = max(0, active.capacity_sarees - sarees_made)

    history = PagdiHistory.objects.filter(employee=employee).order_by("-start_date")

    return render(request, "accounts/pagdi.html", {
        "employee": employee,
        "pagdi": active,
        "sarees_made": sarees_made,
        "remaining": remaining,
        "history": history,
    })


@login_required
def warp_view(request):
    emp_id = request.session.get("employee_id")
    employee = get_object_or_404(Employee, id=emp_id)

    active = WarpHistory.objects.filter(employee=employee, end_date__isnull=True).first()

    sarees_made = 0
    remaining = 0

    if active:
        sarees_made = SareeCount.objects.filter(employee=employee, date__gte=active.start_date).aggregate(Sum("count"))["count__sum"] or 0
        remaining = max(0, active.capacity_sarees - sarees_made)

    history = WarpHistory.objects.filter(employee=employee).order_by("-start_date")

    return render(request, "accounts/warp.html", {
        "employee": employee,
        "warp": active,
        "sarees_made": sarees_made,
        "remaining": remaining,
        "history": history,
    })


@login_required
def employee_history_view(request):
    """
    Combined history page for employee: saree entries, pagdi history, warp history.
    """
    emp = get_object_or_404(Employee, user=request.user)

    saree_history = SareeCount.objects.filter(employee=emp).order_by("-date")
    pagdi_history = PagdiHistory.objects.filter(employee=emp).order_by("-start_date")
    warp_history = WarpHistory.objects.filter(employee=emp).order_by("-start_date")

    return render(request, "accounts/history.html", {
        "employee": emp,
        "saree_history": saree_history,
        "pagdi_history": pagdi_history,
        "warp_history": warp_history,
    })


@login_required
def employee_salary_history(request):
    """
    Employee-facing Salary History: list SalaryHistory rows for this employee.
    """
    emp = get_object_or_404(Employee, user=request.user)
    history = SalaryHistory.objects.filter(employee=emp).order_by("-week_start")
    return render(request, "accounts/employee_salary_history.html", {
        "employee": emp,
        "history": history
    })


# =========================================================
# ADMIN HOME / DASHBOARD
# =========================================================
@staff_required
def admin_home(request):
    today = timezone.localdate()
    monday, sunday = services.get_week_bounds(today)

    total_employees = Employee.objects.count()
    unapproved_count = Employee.objects.filter(is_approved=False).count()
    active_pagdis = PagdiHistory.objects.filter(end_date__isnull=True).count()
    active_warps = WarpHistory.objects.filter(end_date__isnull=True).count()

    stats = {
        "total_employees": total_employees,
        "unapproved_count": unapproved_count,
        "active_pagdis": active_pagdis,
        "active_warps": active_warps,
    }

    return render(request, "accounts/admin/admin_home.html", {
        "stats": stats,
        "week_start": monday,
        "week_end": sunday
    })


@staff_required
def admin_dashboard(request):
    return admin_home(request)


# =========================================================
# ADMIN EMPLOYEES
# =========================================================
@staff_required
def admin_employees(request):
    query = request.GET.get("q", "")

    employees = Employee.objects.all().order_by("-id")
    if query:
        employees = employees.filter(Q(name__icontains=query) | Q(phone__icontains=query))

    return render(request, "accounts/admin/admin_employees.html", {
        "employees": employees,
        "q": query
    })


@staff_required
def admin_employee_detail(request, emp_id):
    """
    Admin detail page for a specific employee.
    Show saree entries (week), pagdi history, warp history and salary history for the employee.
    """
    employee = get_object_or_404(Employee, id=emp_id)
    today = timezone.localdate()
    monday, sunday = services.get_week_bounds(today)

    # weekly entries raw queryset
    weekly_entries_raw = SareeCount.objects.filter(employee=employee, date__gte=monday, date__lte=sunday).order_by("-date")
    # compute salary for each entry (to avoid template arithmetic)
    weekly_entries = []
    rate = employee.salary_per_saree or 0
    for e in weekly_entries_raw:
        weekly_entries.append({
            "id": e.id,
            "date": e.date,
            "count": e.count,
            "notes": e.notes,
            "salary": e.count * rate
        })

    week_sarees = sum([w["count"] for w in weekly_entries]) if weekly_entries else 0
    weekly_salary_before = week_sarees * rate
    week_final = weekly_salary_before - (employee.advance_salary or 0)

    # Employee-level histories
    saree_history_full = SareeCount.objects.filter(employee=employee).order_by("-date")
    pagdi_history_full = PagdiHistory.objects.filter(employee=employee).order_by("-start_date")
    warp_history_full = WarpHistory.objects.filter(employee=employee).order_by("-start_date")
    salary_history_full = SalaryHistory.objects.filter(employee=employee).order_by("-week_start")

    # ---- POST ACTIONS ----
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "approve":
            employee.is_approved = True
            employee.save(update_fields=["is_approved", "updated_at"])
            messages.success(request, "Employee approved.")
            return redirect("admin_employee_detail", emp_id=emp_id)

        if action == "save_salary":
            try:
                new_salary = int(request.POST.get("salary_per_saree"))
                if new_salary < 0:
                    raise ValueError("Negative not allowed")
            except Exception:
                messages.error(request, "Invalid salary value.")
                return redirect("admin_employee_detail", emp_id=emp_id)

            employee.salary_per_saree = new_salary
            employee.save(update_fields=["salary_per_saree", "updated_at"])
            messages.success(request, "Salary updated.")
            return redirect("admin_employee_detail", emp_id=emp_id)

        if action == "add_saree":
            # Admin adding a saree entry for the employee
            try:
                date_str = request.POST.get("date")
                count = int(request.POST.get("count"))
                notes = request.POST.get("notes", "")
            except Exception:
                messages.error(request, "Invalid saree entry.")
                return redirect("admin_employee_detail", emp_id=emp_id)

            SareeCount.objects.create(employee=employee, date=date_str, count=count, notes=notes)
            messages.success(request, "Saree entry added.")
            return redirect("admin_employee_detail", emp_id=emp_id)

        if action == "delete_saree":
            entry_id = request.POST.get("entry_id")
            SareeCount.objects.filter(id=entry_id, employee=employee).delete()
            messages.success(request, "Entry deleted.")
            return redirect("admin_employee_detail", emp_id=emp_id)

    # compute whether week is paid (for UI)
    sh = SalaryHistory.objects.filter(employee=employee, week_start=monday, week_end=sunday).first()
    week_paid = sh.paid_status if sh else False

    return render(request, "accounts/admin/admin_employee_detail.html", {
        "employee": employee,
        "weekly_entries": weekly_entries,
        "weekly_sarees": week_sarees,
        "weekly_salary_before": weekly_salary_before,
        "week_salary": week_final,
        "week_paid": week_paid,
        "week_start": monday,
        "week_end": sunday,
        # full histories for admin view
        "saree_history_full": saree_history_full,
        "pagdi_history_full": pagdi_history_full,
        "warp_history_full": warp_history_full,
        "salary_history_full": salary_history_full,
    })


# =========================================================
# PAGDI / WARP (ADMIN)
# =========================================================
@staff_required
def admin_pagdi_list(request):
    pagdis = PagdiHistory.objects.all().order_by("-start_date")
    result = []

    for p in pagdis.select_related("employee"):
        made = SareeCount.objects.filter(employee=p.employee, date__gte=p.start_date).aggregate(Sum("count"))["count__sum"] or 0
        result.append({
            "obj": p,
            "made": made,
            "remaining": max(0, p.capacity_sarees - made)
        })

    return render(request, "accounts/admin/admin_pagdi_list.html", {"data": result})


@staff_required
def admin_pagdi_create(request):
    employees = Employee.objects.filter(is_approved=True)

    if request.method == "POST":
        emp_id = request.POST.get("employee")
        start = request.POST.get("start_date")
        try:
            capacity = int(request.POST.get("capacity_sarees"))
        except Exception:
            messages.error(request, "Invalid capacity value.")
            return redirect("admin_pagdi_create")

        notes = request.POST.get("notes", "")

        old = PagdiHistory.objects.filter(employee_id=emp_id, end_date__isnull=True).first()
        if old:
            services.finish_pagdi(old.id, request.user, "Auto-finish due to new assignment")

        newp = PagdiHistory.objects.create(employee_id=emp_id, start_date=start, capacity_sarees=capacity, notes=notes)

        PagdiChangeHistory.objects.create(pagdi=newp, employee=newp.employee, admin_user=request.user, action="CREATE", new_capacity=capacity, note=notes)

        messages.success(request, "Pagdi created.")
        return redirect("admin_pagdi_list")

    return render(request, "accounts/admin/admin_pagdi_create.html", {"employees": employees})


@staff_required
def admin_warp_list(request):
    """
    Admin warp listing (similar to pagdi list). Also has Assign Warp button.
    """
    warps = WarpHistory.objects.all().order_by("-start_date")
    result = []
    for w in warps.select_related("employee"):
        made = SareeCount.objects.filter(employee=w.employee, date__gte=w.start_date).aggregate(Sum("count"))["count__sum"] or 0
        result.append({
            "obj": w,
            "made": made,
            "remaining": max(0, w.capacity_sarees - made)
        })
    return render(request, "accounts/admin/admin_warp_list.html", {"data": result})


@staff_required
def admin_warp_create(request):
    employees = Employee.objects.filter(is_approved=True)

    if request.method == "POST":
        WarpHistory.objects.create(employee_id=request.POST.get("employee"), capacity_sarees=int(request.POST.get("capacity") or 0), start_date=timezone.localdate())
        messages.success(request, "Warp assigned.")
        return redirect("admin_warp_list")

    return render(request, "accounts/admin/admin_warp_create.html", {"employees": employees})


# =========================================================
# WEEKLY SALARY (ADMIN)
# =========================================================
@staff_required
def admin_weekly_salary(request):
    today = timezone.localdate()
    monday, sunday = services.get_week_bounds(today)

    rows = []
    employees = Employee.objects.all()

    for emp in employees:
        sarees = SareeCount.objects.filter(employee=emp, date__gte=monday, date__lte=sunday).aggregate(Sum("count"))["count__sum"] or 0
        salary_before = sarees * (emp.salary_per_saree or 0)
        advance_amt = emp.advance_salary or 0
        final_salary = salary_before - advance_amt

        sh = SalaryHistory.objects.filter(employee=emp, week_start=monday, week_end=sunday).first()
        paid = sh.paid_status if sh else False

        rows.append({
            "employee": emp,
            "sarees": sarees,
            "salary_before": salary_before,
            "advance": advance_amt,
            "final_salary": final_salary,
            "paid": paid,
        })

    return render(request, "accounts/admin/admin_weekly_salary.html", {"rows": rows, "week_start": monday, "week_end": sunday})


@staff_required
def admin_salary_history(request):
    history = SalaryHistory.objects.select_related("employee").all().order_by("-week_start")
    return render(request, "accounts/admin/admin_salary_history.html", {"history": history})


# =========================================================
# SALARY ACTIONS (ADMIN)
# =========================================================
@staff_required
def give_advance_view(request, emp_id):
    if request.method != "POST":
        return HttpResponseBadRequest("POST only")
    try:
        amount = int(request.POST.get("amount"))
    except Exception:
        return HttpResponseBadRequest("Invalid amount")
    note = request.POST.get("note", "")
    with transaction.atomic():
        services.give_advance(emp_id, amount, request.user, note)
    messages.success(request, f"Advance ₹{amount} added.")
    return redirect("admin_weekly_salary")


@staff_required
def clear_advance(request, emp_id):
    if request.method != "POST":
        return HttpResponseBadRequest("POST only")
    services.clear_advance_for_employee(emp_id, request.user, "Cleared by admin")
    messages.success(request, "Advance cleared.")
    return redirect("admin_weekly_salary")


@staff_required
def mark_paid(request, emp_id):
    if request.method != "POST":
        return HttpResponseBadRequest("POST only")
    emp = get_object_or_404(Employee, id=emp_id)
    today = timezone.localdate()
    monday, sunday = services.get_week_bounds(today)

    sarees = SareeCount.objects.filter(employee=emp, date__gte=monday, date__lte=sunday).aggregate(Sum("count"))["count__sum"] or 0
    rate = emp.salary_per_saree or 0
    total = sarees * rate
    advance = emp.advance_salary or 0
    final = total - advance
    note = request.POST.get("note", "")

    sh, created = SalaryHistory.objects.get_or_create(employee=emp, week_start=monday, week_end=sunday, defaults={
        "sarees": sarees,
        "salary_rate": rate,
        "total_salary_before_advance": total,
        "advance_salary": advance,
        "final_salary": final,
        "paid_status": True,
        "paid_date": today,
        "notes": note or "Paid"
    })

    if not created:
        sh.paid_status = True
        sh.paid_date = today
        sh.final_salary = final
        sh.notes = note or sh.notes
        sh.save(update_fields=["paid_status", "paid_date", "final_salary", "notes"])


    messages.success(request, "Marked paid.")
    return redirect("admin_weekly_salary")


@staff_required
def mark_unpaid(request, emp_id):
    if request.method != "POST":
        return HttpResponseBadRequest("POST only")
    emp = get_object_or_404(Employee, id=emp_id)
    today = timezone.localdate()
    monday, sunday = services.get_week_bounds(today)
    sh = SalaryHistory.objects.filter(employee=emp, week_start=monday, week_end=sunday).first()
    if sh:
        sh.paid_status = False
        sh.paid_date = None
        sh.save(update_fields=["paid_status", "paid_date"])

    messages.success(request, "Marked unpaid.")
    return redirect("admin_weekly_salary")


# =========================================================
# PDF EXPORT (single employee salary slip)
# =========================================================
@staff_required
def salary_slip_pdf(request, emp_id):
    import io
    from reportlab.pdfgen import canvas

    emp = get_object_or_404(Employee, id=emp_id)
    today = timezone.localdate()
    monday, sunday = services.get_week_bounds(today)

    sarees = SareeCount.objects.filter(employee=emp, date__gte=monday, date__lte=sunday).aggregate(Sum("count"))["count__sum"] or 0
    final = (sarees * (emp.salary_per_saree or 0)) - (emp.advance_salary or 0)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="salary_slip_{emp.name}.pdf"'

    p = canvas.Canvas(response)
    p.drawString(100, 800, f"Salary Slip for {emp.name}")
    p.drawString(100, 780, f"Week: {monday} — {sunday}")
    p.drawString(100, 760, f"Final Salary: ₹{final}")
    p.showPage()
    p.save()
    return response


# =========================================================
# ADMIN: SAREE ENTRY & APPROVE
# =========================================================
@staff_required
def admin_saree_entry(request):
    employees = Employee.objects.filter(is_approved=True).order_by("name")
    if request.method == "POST":
        emp_id = request.POST.get("employee")
        try:
            count = int(request.POST.get("count"))
        except Exception:
            messages.error(request, "Invalid count")
            return redirect("admin_saree_entry")
        date = request.POST.get("date") or timezone.localdate()
        notes = request.POST.get("notes", "")
        SareeCount.objects.create(employee_id=emp_id, count=count, date=date, notes=notes)
        messages.success(request, "Saree entry added.")
        return redirect("admin_saree_entry")
    return render(request, "accounts/admin/admin_saree_entry.html", {"employees": employees})


@staff_required
def admin_approve_employee(request, emp_id):
    emp = get_object_or_404(Employee, id=emp_id)
    emp.is_approved = True
    emp.save(update_fields=["is_approved", "updated_at"])
    messages.success(request, f"{emp.name} approved.")
    return redirect("admin_employees")


# =========================================================
# EXCEL / GLOBAL HISTORY DOWNLOAD (XLSX)
# =========================================================
@staff_required
def download_global_history(request):
    """
    Exports all history data (Saree, Pagdi, Warp, Salary) into a single Excel file.
    """
    wb = openpyxl.Workbook()

    # Sheet 1 - Saree History
    ws = wb.active
    ws.title = "Saree History"
    headers = ["Employee", "Date", "Count", "Notes", "Salary Earned"]
    ws.append(headers)
    for h in SareeCount.objects.select_related("employee").order_by("-date"):
        salary = h.count * (h.employee.salary_per_saree or 0)
        ws.append([h.employee.name, str(h.date), h.count, h.notes or "", salary])
    for col in range(1, len(headers) + 1):
        ws[f"{get_column_letter(col)}1"].font = Font(bold=True)

    # Sheet 2 - Pagdi History
    ws2 = wb.create_sheet("Pagdi History")
    headers = ["Employee", "Start", "End", "Capacity", "Notes"]
    ws2.append(headers)
    for p in PagdiHistory.objects.select_related("employee").order_by("-start_date"):
        ws2.append([p.employee.name, str(p.start_date), str(p.end_date) if p.end_date else "Active", p.capacity_sarees, p.notes or ""])
    for col in range(1, len(headers) + 1):
        ws2[f"{get_column_letter(col)}1"].font = Font(bold=True)

    # Sheet 3 - Warp History
    ws3 = wb.create_sheet("Warp History")
    headers = ["Employee", "Start", "End", "Capacity", "Notes"]
    ws3.append(headers)
    for w in WarpHistory.objects.select_related("employee").order_by("-start_date"):
        ws3.append([w.employee.name, str(w.start_date), str(w.end_date) if w.end_date else "Active", w.capacity_sarees, w.notes or ""])
    for col in range(1, len(headers) + 1):
        ws3[f"{get_column_letter(col)}1"].font = Font(bold=True)

    # Sheet 4 - Salary History
    ws4 = wb.create_sheet("Salary History")
    headers = ["Employee", "Week Start", "Week End", "Sarees", "Rate", "Advance", "Final", "Paid", "Notes"]
    ws4.append(headers)
    for s in SalaryHistory.objects.select_related("employee").order_by("-week_start"):
        ws4.append([s.employee.name, str(s.week_start), str(s.week_end), s.sarees, s.salary_rate, s.advance_salary, s.final_salary, "Yes" if s.paid_status else "No", s.notes or ""])
    for col in range(1, len(headers) + 1):
        ws4[f"{get_column_letter(col)}1"].font = Font(bold=True)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = "Global_History_Report.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@staff_required
def download_global_weekly_salary(request):
    """
    Export ALL salary history weeks (past + present) into XLSX.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Salary History"

    headers = ["Employee", "Week Start", "Week End", "Sarees", "Rate", "Advance", "Final", "Paid?", "Notes"]
    ws.append(headers)

    # Bold headers
    for col in range(1, len(headers) + 1):
        ws[f"{get_column_letter(col)}1"].font = Font(bold=True)

    history = SalaryHistory.objects.select_related("employee").order_by("-week_start")

    for s in history:
        ws.append([
            s.employee.name,
            str(s.week_start),
            str(s.week_end),
            s.sarees,
            s.salary_rate,
            s.advance_salary,
            s.final_salary,
            "Yes" if s.paid_status else "No",
            s.notes or ""
        ])

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Global_Weekly_Salary.xlsx"'

    wb.save(response)
    return response
